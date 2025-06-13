import io
import openpyxl
from rest_framework import renderers

# Define o Cabeçalho dos arquivos em listas de valores
SENSOR_DATA_FILE_HEADERS = ["id","sensor","mac_address","unidade_med","latitude","longitude","status"]
AMBIENTE_DATA_FILE_HEADERS = ["id","sig","descricao","ni","responsavel"]
HISTORICO_DATA_FILE_HEADERS = ["id","valor","timestamp","sensor","ambiente"]

# Cria arquivo excel listando dados do SENSOR a partir de tabela com cabeçalho definido em SENSOR_DATA_FILE_HEADERS
class ExcelSensorDataRenderer(renderers.BaseRenderer):

    # Cria um arquivo excel base
    media_type = "application/vnd.ms-excel"
    format = "xlsx"

    # Recebe dados do construtor
    def render(self, data, accepted_media_type=None, renderer_context=None):    

        # Cria um ambiente de trabalho(planilha sem extensão de arquivo) com um buffer(Ambiente que "carrega" os dados antes de salvar/commit)
        workbook = openpyxl.Workbook()
        buffer = io.BytesIO()
        worksheet = workbook.active
        # Insere o Cabeçalho da planilha, definido em SENSOR_DATA_FILE_HEADERS
        worksheet.append(SENSOR_DATA_FILE_HEADERS)

        # Le linha a linha dos dados do construtor, e insere no final da planilha
        for sensor_data in data:
            worksheet.append(list(sensor_data.values()))
        # Salva os dados do buffer
        workbook.save(buffer)
        
        # Devolve o buffer
        return buffer.getvalue()
    
# Cria arquivo excel listando dados do AMBIENTE a partir de tabela com cabeçalho definido em AMBIENTE_DATA_FILE_HEADERS
class ExcelAmbienteDataRenderer(renderers.BaseRenderer):
    
    # Cria um arquivo excel base
    media_type = "application/vnd.ms-excel"
    format = "xlsx"
    
    # Recebe dados do construtor
    def render(self, data, accepted_media_type=None, renderer_context=None):    
        # Cria um ambiente de trabalho(planilha sem extensão de arquivo) com um buffer(Ambiente que "carrega" os dados antes de salvar/commit)
        workbook = openpyxl.Workbook()
        buffer = io.BytesIO()
        worksheet = workbook.active
        # Insere o Cabeçalho da planilha, definido em AMBIENTE_DATA_FILE_HEADERS
        worksheet.append(AMBIENTE_DATA_FILE_HEADERS)

        # Le linha a linha dos dados do construtor, e insere no final da planilha
        for ambiente_data in data:
            worksheet.append(list(ambiente_data.values()))
        # Salva os dados do buffer
        workbook.save(buffer)
        
        # Devolve o buffer
        return buffer.getvalue()

# Cria arquivo excel listando dados do HISTORICO a partir de tabela com cabeçalho definido em HISTORICO_DATA_FILE_HEADERS    
class ExcelHistoricoDataRenderer(renderers.BaseRenderer):
    
    # Cria um arquivo excel base
    media_type = "application/vnd.ms-excel"
    format = "xlsx"

    # Recebe dados do construtor
    def render(self, data, accepted_media_type=None, renderer_context=None):    
        
        # Cria um ambiente de trabalho(planilha sem extensão de arquivo) com um buffer(Ambiente que "carrega" os dados antes de salvar/commit)
        workbook = openpyxl.Workbook()
        buffer = io.BytesIO()
        worksheet = workbook.active
        # Insere o Cabeçalho da planilha, definido em HISTORICO_DATA_FILE_HEADERS
        worksheet.append(HISTORICO_DATA_FILE_HEADERS)

        # Le linha a linha dos dados do construtor, e insere no final da planilha
        for historico_data in data:
            worksheet.append(list(historico_data.values()))
        # Salva os dados do buffer
        workbook.save(buffer)

        # Devolve o buffer
        return buffer.getvalue()

